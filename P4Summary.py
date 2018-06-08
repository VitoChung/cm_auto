import const
from P4 import P4, P4Exception  # Import the module
import re

import sys
import time
import gspread
from oauth2client.service_account import ServiceAccountCredentials as SAC

import openpyxl
from openpyxl.utils import get_column_letter

import requests
from requests_ntlm import HttpNtlmAuth


def main():
    const.trend_ishare_upload = False
    for tmp_code_line in const.code_lines:
        try:
            const.code_line = tmp_code_line
            print(const.code_line.google_sheet_name)
            readme = {}
            if const.code_line.trend_readme_url != '' and const.trend_password != '':
                html = login_hotfix_portal()
                readme = parse_hotifx_readme(html)
            hotfixes = parse_P4_description(readme)
            update_spreadsheet(hotfixes, readme)

            if const.trend_password != '':
                download_sharepoint()
                update_excel_file(hotfixes, readme)
                if const.trend_ishare_upload == True:
                    upload_sharepoint()
                    const.trend_ishare_upload = False

            print(const.code_line.google_sheet_name + ' updated...')
            print('')
        except:
            print("Unexpected error:", sys.exc_info()[0])


def parse_P4_description(readme):
    print('Get P4 hotfix description...')

    p4 = P4()  # Create the P4 instance
    p4.port = const.p4_port
    p4.user = const.p4_user
    p4.client = const.p4_client  # Set some environment variables
    p4.set_env("P4PASSWD", const.p4_password)

    try:  # Catch exceptions with try/except
        p4.connect()  # Connect to the Perforce server
        p4.run_login()

        hotfixes = []
        change_history = p4.run('changes', '-l', const.code_line.p4_path_QA + '...')
        for change in change_history:
            change_no = change.get('change')
            change_time = time.strftime("%Y/%m/%d %H:%M:%S", time.localtime(int(change.get('time'))))
            chnage_user = change.get('user')
            change_desc = change.get('desc')
            change_file = ''

            desc_user = ''
            desc_case = ''
            desc_hotfix = ''
            desc_build = ''

            description_list = re.split(' |,|\.|\(|\)|\?', change_desc)
            for description in description_list:
                tmp_desc = str(description).strip().upper()
                if 'TT' in tmp_desc or 'SEG' in tmp_desc or 'VRTS' in tmp_desc:
                    desc_case += refine_QA_changelist(tmp_desc) + '\n'
                elif 'HF' in tmp_desc:
                    # desc_hotfix = tmp_desc
                    desc_hotfix = 'HFB' + get_hotfix_number(tmp_desc)
                elif 'B' in tmp_desc and 'BY' not in tmp_desc and 'BACK' not in tmp_desc and 'BASIC' not in tmp_desc:
                    desc_build = tmp_desc
                elif desc_user == '' and tmp_desc != '':
                    tmp_user = const.dictQA.get(tmp_desc)
                    if tmp_user != None:
                        desc_user = tmp_user
            desc_case = desc_case.rstrip('\n')

            if desc_user == '':
                desc_user = chnage_user
            if desc_build == '':
                desc_build = 'B'

            # get hotfix fileList
            file_list = p4.run('describe', change_no)
            for file in file_list[0].get('depotFile'):
                tmp_file = str(file).replace(const.code_line.p4_path_QA, '')
                if tmp_file != '' and 'Hotfix.ini'.upper() not in tmp_file.upper():
                    change_file += tmp_file[1 + tmp_file.find('/'):] + '\n'

            change_file = change_file.rstrip('\n')
            if len(change_file) > 50000:
                # print(change_file)
                change_file = 'Exceed 50,000 characters'

            # end ##########

            # get RD changeList
            change_RD_list = []
            change_RD_history = p4.run('changes', '-l',
                                       const.code_line.p4_path_RD + '...@' + const.p4_start_date + ',@now')
            for change_RD in change_RD_history:
                if change_RD.get('user') not in const.dictQA.values():
                    change_RD_no = change_RD.get('change')
                    change_RD_desc = change_RD.get('desc')
                    if 'TT' in change_RD_desc or 'SEG' in change_RD_desc or 'VRTS' in change_RD_desc:
                        change_RD_list.append([change_RD_no, refine_RD_changelist(change_RD_desc)])

            change_list_for_hotfix = ''
            for case in desc_case.split('\n'):
                if case == '':
                    continue

                if 'SEG' in case or 'VRTS' in case:
                    case_ID = 'Jira-' + case
                else:
                    case_ID = 'US-SEG-' + case.replace('TT', '')

                change_list_for_case = ''
                fixes = p4.run('fixes', '-j', case_ID)
                if len(fixes) > 0:
                    for fix in fixes:
                        change_list_for_case += fix.get('Change') + '\n'
                else:
                    for list in change_RD_list:
                        if case in list[1]:
                            change_list_for_case += list[0] + '\n'

                if change_list_for_case.rstrip('\n') != '':
                    change_list_for_hotfix += '[' + case + ']' + '\n' + change_list_for_case

            change_list_for_hotfix = change_list_for_hotfix.rstrip('\n')
            # end ##########
            hotfix_readme = readme.get(get_hotfix_number(desc_hotfix))

            hotfixes.append([const.code_line.lang, desc_hotfix, desc_case, hotfix_readme, desc_build, change_no,
                             change_list_for_hotfix, change_file, chnage_user, desc_user, change_time])

        hotfixes.reverse()
        return hotfixes

    except P4Exception:
        print('Exception!!!')
        for e in p4.errors:
            print(e)
    except:
        print("Unexpected error:", sys.exc_info()[0])
    finally:
        p4.disconnect()


def update_spreadsheet(hotfixes, readme):
    print('Update Spreadsheet...')

    scope = [const.google_sheet_url]
    credentials = SAC.from_json_keyfile_name(const.google_json_file, scope)
    gc = gspread.authorize(credentials)
    worksheet = gc.open_by_key(const.google_sheet_key).worksheet(const.code_line.google_sheet_name)

    next_data = exist_hotfix(worksheet)

    row_count = next_data[0]
    for hotfix in hotfixes:
        if time.strptime(hotfix[10], '%Y/%m/%d %H:%M:%S') > time.strptime(next_data[1], '%Y/%m/%d %H:%M:%S'):
            col_count = 1
            for o in hotfix:
                worksheet.update_cell(row_count, col_count, o)
                col_count += 1
            row_count += 1

    # update empty reademe
    for empty_row_count in range(2, row_count):
        readme_hotfix_number = get_hotfix_number(worksheet.cell(empty_row_count, 2).value)
        if (worksheet.cell(empty_row_count, 4).value == 'None' or worksheet.cell(empty_row_count, 4).value == '') and (
                worksheet.cell(empty_row_count, 2).value != '') and \
                (str(readme.get(readme_hotfix_number)) != 'None' and str(readme.get(readme_hotfix_number)) != ''):
            worksheet.update_cell(empty_row_count, 4, readme.get(readme_hotfix_number))

    export_spreadsheet(worksheet)


def export_spreadsheet(worksheet):
    export_file = worksheet.export(format='xlsx')
    f = open('filename.xlsx', 'wb')
    f.write(export_file)
    f.close()


def login_hotfix_portal():
    print('Login Portal...')

    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    usernameStr = const.trend_domain_account
    passwordStr = const.trend_password

    browser = webdriver.Chrome()
    try:
        browser.get((const.code_line.trend_readme_url))
        # fill in username and hit the next button
        username = browser.find_element_by_id('userNameInput')
        username.send_keys(usernameStr)

        # wait for transition then continue to fill items
        password = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.ID, 'passwordInput')))
        password.send_keys(passwordStr)

        signInButton = browser.find_element_by_id('submitButton')
        signInButton.click()

        return browser.page_source
    except:
        browser.close()
        return None


def parse_hotifx_readme(html):
    print('Get Hotfix Readme...')
    try:
        from html.parser import HTMLParser

        class MyHTMLParser(HTMLParser):
            def __init__(self):
                HTMLParser.__init__(self)
                self.data = []

            def handle_data(self, data):
                if str(data).strip() != '':
                    self.data.append(data)

        parser = MyHTMLParser()
        parser.feed(html)
        content = ''
        hotfix_no = ''
        readme = {}
        for line in parser.data:
            if 'Patch Build' in line or 'Hotfix Build' in line or 'GM Build' in line:

                if hotfix_no != '':
                    readme[hotfix_no] = content.strip()
                    hotfix_no = ''
                    content = ''

                hotfix_no = str(line).split()[-1]
            else:
                if hotfix_no != '':
                    content += str(line).strip() + '\n'

        if content.strip() != '':
            readme[hotfix_no] = content.strip()

        return readme
    except:
        print("Unexpected error:", sys.exc_info()[0])
    # finally:
    #   p4.disconnect()


def update_excel_file(hotfixes, readme):
    print('Update Excel File...')
    wb = openpyxl.load_workbook(const.trend_ishare_download_file)

    sheets = wb.get_sheet_names()
    sheet0 = sheets[0]
    ws = wb.get_sheet_by_name(sheet0)

    last_row = ws.max_row
    latest_hotfix_number = str(ws['{}{}'.format('A', last_row)].value).split('(')[-1].replace(')', '')

    row_count = last_row + 1
    for hotfix in hotfixes:
        if get_hotfix_number(hotfix[1]) > latest_hotfix_number:
            const.trend_ishare_upload = True
            col_count = 1

            ws['{}{}'.format('A', row_count)] = hotfix[2]
            ws['{}{}'.format('C', row_count)] = hotfix[3]
            ws['{}{}'.format('E', row_count)] = hotfix[4]
            ws['{}{}'.format('F', row_count)] = hotfix[7]
            ws['{}{}'.format('G', row_count)] = hotfix[6]
            ws['{}{}'.format('H', row_count)] = hotfix[10].split(' ')[0]
            ws['{}{}'.format('I', row_count)] = 'No'
            copy_row_format(ws, row_count, last_row - 1)
            row_count += 1

            if '6.0' in const.code_line.google_sheet_name:
                ws['{}{}'.format('A', row_count)] = 'TMCM 6.0 Hot Fix (' + get_hotfix_number(hotfix[1]) + ')'
            else:
                ws['{}{}'.format('A', row_count)] = 'TMCM 7.0 Hot Fix (' + get_hotfix_number(hotfix[1]) + ')'
            ws['{}{}'.format('H', row_count)] = hotfix[9]
            copy_row_format(ws, row_count, last_row)
            row_count += 1

    wb.save(const.trend_ishare_upload_file)


def download_sharepoint():
    print('Download Excel from iShare...')
    session = requests.Session()
    session.auth = HttpNtlmAuth(const.trend_domain_account, const.trend_password, session)

    resp = requests.get(const.code_line.trend_ishare_url, auth=session.auth)
    if resp.ok:
        with open(const.trend_ishare_download_file, 'wb') as file:
            file.write(resp.content)


def upload_sharepoint():
    print('Upload Excel to iShare...')
    session = requests.Session()
    session.auth = HttpNtlmAuth(const.trend_domain_account, const.trend_password, session)

    file = open(const.trend_ishare_upload_file, 'rb')
    bytes = bytearray(file.read())
    resp = requests.put(const.code_line.trend_ishare_url, data=bytes, auth=session.auth)
    # print(resp.status_code)


def refine_QA_changelist(desc):
    tmp_desc = str(desc).strip().upper()
    tmp_desc = tmp_desc.replace('TT-', 'TT')
    tmp_desc = tmp_desc.replace('SEG', 'SEG-')
    tmp_desc = tmp_desc.replace('SEG--', 'SEG-')
    return tmp_desc


def refine_RD_changelist(desc):
    tmp_desc = str(desc).strip().upper()
    tmp_desc = tmp_desc.replace('TT-', 'TT')

    tmp_desc = tmp_desc.replace('SEG_', 'SEG-')
    tmp_desc = tmp_desc.replace('SEGTT-', 'SEG-')
    tmp_desc = tmp_desc.replace('SEGTT:', 'SEG-')
    tmp_desc = tmp_desc.replace('SEG CASE - ', 'SEG-')
    tmp_desc = tmp_desc.replace('SEG--', 'SEG-')

    tmp_desc = tmp_desc.replace('SEG-VRTS', 'VRTS-')
    tmp_desc = tmp_desc.replace('SEGTT-VRTS-', 'VRTS-')
    tmp_desc = tmp_desc.replace('VRTS--', 'VRTS-')
    return tmp_desc


def exist_hotfix(worksheet):
    submit_datetime_list = worksheet.col_values(11)
    submit_datetime_list = [k for k in submit_datetime_list if k != '']
    next_row = len(submit_datetime_list) + 1

    if next_row > 2:
        submit_datetime_list = sorted(submit_datetime_list, reverse=True)
        return [next_row, submit_datetime_list[1]]
    else:
        return [next_row, '1970/01/01 00:00:00']


def get_hotfix_number(hotfix):
    return str(hotfix).strip().replace('HF', '').replace('B', '')


def copy_cell_format(target, source):
    target.number_format = source.number_format
    target.font = source.font.copy()
    target.alignment = source.alignment.copy()
    target.border = source.border.copy()
    target.fill = source.fill.copy()


def copy_row_format(ws, target, source):
    for col_count in range(1, ws.max_column + 1):
        copy_cell_format(ws['{}{}'.format(get_column_letter(col_count), target)],
                         ws['{}{}'.format(get_column_letter(col_count), source)])


if __name__ == "__main__":
    main()
